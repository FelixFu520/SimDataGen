#!/usr/bin/env python3
"""
从 workdir_omni(或任意同结构根目录)按「场景」聚合任务目录,每个场景随机选一个 seed 子目录,
取 rgb/CAM_A 下字典序第一帧,将 CAM_A/B/C/D 的 rgb、depth、semantic_vis 嵌入 Excel。

pip install openpyxl pillow   # 若尚未安装
python tools/export/export_omni_samples_excel.py
python tools/export/export_omni_samples_excel.py /path/to/workdir_omni -o samples.xlsx --seed 42
"""

from __future__ import annotations

import argparse
import random
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# 任务目录名：{场景}_{seed}_{A}_{B} ,seed 为后缀 _A_B 前紧邻的一段数字(如 _5_100、_40_10)
TASK_DIR_SUFFIX_RE = re.compile(r"^(.+)_(\d+)_(\d+_\d+)$")

IMAGE_SUFFIXES = (".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff", ".bmp")
CAMS = ("CAM_A", "CAM_B", "CAM_C", "CAM_D")


def repo_root() -> Path:
    return Path(__file__).resolve().parent.parent.parent


def parse_task_dir(name: str) -> tuple[str, int] | None:
    m = TASK_DIR_SUFFIX_RE.match(name)
    if not m:
        return None
    return m.group(1), int(m.group(2))


def group_scenes(root: Path) -> dict[str, list[Path]]:
    scenes: dict[str, list[Path]] = {}
    for p in root.iterdir():
        if not p.is_dir():
            continue
        parsed = parse_task_dir(p.name)
        if not parsed:
            continue
        scene, _seed = parsed
        scenes.setdefault(scene, []).append(p)
    for paths in scenes.values():
        paths.sort(key=lambda x: x.name)
    return scenes


def first_frame_stem(task_dir: Path) -> str | None:
    cam_a = task_dir / "rgb" / "CAM_A"
    if not cam_a.is_dir():
        return None
    files: list[Path] = []
    for p in cam_a.iterdir():
        if p.is_file() and p.suffix.lower() in IMAGE_SUFFIXES:
            files.append(p)
    if not files:
        return None
    files.sort(key=lambda p: p.name)
    return files[0].stem


def find_image(modality_dir: Path, stem: str) -> Path | None:
    """在 modality_dir(已是 CAM_X 目录)下按 stem 找一张图。"""
    if not modality_dir.is_dir():
        return None
    for suf in IMAGE_SUFFIXES:
        cand = modality_dir / f"{stem}{suf}"
        if cand.is_file():
            return cand
    # 兜底：stem 前缀匹配(极少见)
    for p in modality_dir.iterdir():
        if p.is_file() and p.stem == stem and p.suffix.lower() in IMAGE_SUFFIXES:
            return p
    return None


def add_sheet_thumbnail(ws, row: int, col: int, image_path: Path, *, max_side: int) -> None:
    img = XLImage(str(image_path))
    if img.width and img.height:
        w, h = float(img.width), float(img.height)
        scale = min(max_side / w, max_side / h, 1.0)
        img.width = int(w * scale)
        img.height = int(h * scale)
    cell = f"{get_column_letter(col)}{row}"
    ws.add_image(img, cell)


def build_workbook(
    root: Path,
    *,
    rng: random.Random,
    thumb_max: int,
) -> tuple[Workbook, list[str]]:
    scenes = group_scenes(root)
    if not scenes:
        raise SystemExit(
            f"未在目录中找到符合命名规则 *_<seed>_<A>_<B> 的子文件夹(如 *_75_5_100): {root}"
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "samples"

    # 表头两行：第一行合并分组,第二行 CAM
    ws["A1"] = "场景名称"
    ws["B1"] = "seed"
    ws.merge_cells("C1:F1")
    ws["C1"] = "rgb(四路)"
    ws.merge_cells("G1:J1")
    ws["G1"] = "depth(四路)"
    ws.merge_cells("K1:N1")
    ws["K1"] = "semantic_vis(四路)"
    for addr in ("C1", "G1", "K1"):
        ws[addr].alignment = Alignment(horizontal="center", vertical="center")

    for c in range(3, 15):
        ws.cell(row=2, column=c, value=CAMS[(c - 3) % 4])

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    img_row_height = max(90.0, thumb_max * 0.75)
    warnings: list[str] = []

    data_row = 3
    for scene in sorted(scenes.keys()):
        task_dirs = scenes[scene]
        chosen = rng.choice(task_dirs)
        parsed = parse_task_dir(chosen.name)
        assert parsed is not None
        _scene_name, seed_val = parsed
        stem = first_frame_stem(chosen)
        if stem is None:
            warnings.append(f"[跳过] {scene}: 随机目录无 rgb/CAM_A 首帧: {chosen.name}")
            continue

        blocks = [
            ("rgb", "rgb"),
            ("depth", "depth"),
            ("semantic_vis", "semantic_vis"),
        ]
        paths_flat: list[Path] = []
        missing: str | None = None
        for _label, subdir in blocks:
            for cam in CAMS:
                rel = chosen / subdir / cam
                pth = find_image(rel, stem)
                if not pth:
                    missing = f"{subdir}/{cam}/{stem}.*"
                    break
                paths_flat.append(pth)
            if missing:
                break
        if missing:
            warnings.append(f"[跳过] {scene} seed={seed_val} 缺图: {missing}")
            continue

        ws.cell(row=data_row, column=1, value=scene)
        ws.cell(row=data_row, column=2, value=seed_val)
        ws.row_dimensions[data_row].height = img_row_height

        col = 3
        for pth in paths_flat:
            add_sheet_thumbnail(ws, data_row, col, pth, max_side=thumb_max)
            col += 1

        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 10
        for c in range(3, 15):
            ws.column_dimensions[get_column_letter(c)].width = thumb_max / 5.5 + 2

        data_row += 1

    if data_row == 3:
        raise SystemExit("没有成功写入任何场景行(可能全部缺图或缺首帧),见上述警告。")

    return wb, warnings


def main() -> None:
    ap = argparse.ArgumentParser(description="按场景随机 seed 抽样首帧,导出多路相机图到 Excel。")
    ap.add_argument(
        "workdir",
        nargs="?",
        default=str(repo_root() / "workdir_omni"),
        type=Path,
        help="数据根目录(默认：仓库下 workdir_omni)",
    )
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=repo_root()  / "omni_scene_samples.xlsx",
        help="输出 xlsx 路径(默认 tools/omni_scene_samples.xlsx)",
    )
    ap.add_argument("--seed", type=int, default=0, help="随机种子(可复现抽样)")
    ap.add_argument(
        "--thumb",
        type=int,
        default=160,
        help="缩略图最长边像素(默认 160,越大文件越大)",
    )
    args = ap.parse_args()

    root = args.workdir.expanduser().resolve()
    if not root.is_dir():
        raise SystemExit(f"目录不存在: {root}")

    rng = random.Random(args.seed)
    wb, warns = build_workbook(root, rng=rng, thumb_max=args.thumb)
    out = args.output.expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"已写入: {out}")
    for w in warns[:50]:
        print(w)
    if len(warns) > 50:
        print(f"... 另有 {len(warns) - 50} 条警告未显示")


if __name__ == "__main__":
    main()
